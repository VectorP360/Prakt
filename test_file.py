from datetime import datetime
import os
from typing import Optional

from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from transliterate import translit

from enum_class import StrChoice
from repository.user_repository import UserIn
from repository.facility_repository import FacilityRepository
from repository.posts_repository import PostsRepository
from password_generator import password_generator
from issue_classes.add_issue.facility_issue import FacilityIssue
from issue_classes.add_issue.fathersname_issue import FathresnameIssue
from issue_classes.add_issue.hire_date_issue import HireDateIssue
from issue_classes.add_issue.name_issue import NameIssue
from issue_classes.add_issue.post_issue import PostIssue
from issue_classes.add_issue.surname_issue import SurnameIssue

def read_from_excel(posts_repository: PostsRepository, facility_repository : FacilityRepository) -> Optional[list]:

    name = NameIssue()
    surname = SurnameIssue()
    fathersname = FathresnameIssue()
    hire_date = HireDateIssue()
    post = PostIssue()
    facility = FacilityIssue()

    name.set_next(surname).set_next(fathersname).set_next(hire_date).set_next(post).set_next(facility)

    try:
        excel_file = str(input('Введите путь до требуемого excel файла: '))
        workbook = load_workbook(excel_file)    
    except FileNotFoundError:
        print('Указанный файл не найден')
        input('Нажмите Enter что бы продолжить')
        return
    except InvalidFileException:
        print('Введён файл не подходящего формата!')
        input('Нажмите Enter что бы продолжить')
        return
    sheet = workbook.active
    data = []
    
    row = sheet[1]
    if not name.handle(row):
        print("НЕПРАВИЛЬНО")
        return

    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            final_date = row[3]
            if type(row[3]) != datetime:
                final_date = datetime.today()

            fetched_post = posts_repository.get_by_name(row[4])

            fetched_facility = facility_repository.get_by_name(row[5])

            login = translit(row[0][:3] + row[1][:3] + row[2][:3],"ru", reversed="true")

            match fetched_post.name:
                case StrChoice.gendir_post:
                    password_length = 12
                case StrChoice.glavspec_post:
                    password_length = 12
                case StrChoice.starspec_post:
                    password_length = 8
                case StrChoice.spec1categ_post:
                    password_length = 8

            password = password_generator(password_length)

            data.append([row[0],row[1],row[2],fetched_facility.name,fetched_post.name,final_date,login,password])
        except IndexError:
            print('В таблице недостаточно столбцов')
    
    return data
    


def create_raport(data: list):
    if not os.path.exists('./raports'):
        os.mkdir('./raports')

    raportname = str('./raports/raport' + str(datetime.date(datetime.today())) + '.xlsx')

    if not os.path.exists(raportname):
        raport = Workbook()
        raport_sheet = raport.active
        raport_sheet['A1'] = 'ФИО'
        raport_sheet['B1'] = 'Установка'
        raport_sheet['C1'] = 'Логин'
        raport_sheet['D1'] = 'Пароль'
        raport.save(filename=raportname)

    raport = load_workbook(raportname)
    raport_sheet = raport['Sheet']

    row_number: int = 2

    for iteration in data:

        raport_sheet [f'A{row_number}'] = (f'{iteration[1]} {iteration[0]} {iteration[2]}')
        raport_sheet [f'B{row_number}'] = iteration[3]
        raport_sheet [f'C{row_number}'] = iteration[6]
        raport_sheet [f'D{row_number}'] = iteration[7]
        raport.save(raportname)
        row_number +=1
    raport.close()
