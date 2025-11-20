from datetime import datetime
import os
from typing import Optional, List, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from transliterate import translit

from src.enums import UserPost
from src.tools.password_generator import PasswordGenerator
from src.repositories.repository import RepositoryManager
from src.handlers import (
    FacilityHandler,
    FathresnameHandler,
    HireDateHandler,
    NameHandler,
    PostHandler,
    SurnameHandler,
)
from src.handlers.type_handlers import (
    TypeNameHandler,
    TypeSurnameHandler,
    TypeFathresnameHandler,
    TypeHireDateHandler,
    TypeFacilityHandler,
    TypePostHandler,
)


class NewUsersExcelReader:
    def __init__(self, repository_manager: RepositoryManager) -> None:
        self.repository_manager = repository_manager

    def __check_file(self, workbook: Workbook) -> Optional[Worksheet]:
        name_handler = NameHandler()
        surname_handler = SurnameHandler()
        fathersname_handler = FathresnameHandler()
        hire_date_handler = HireDateHandler()
        post_handler = PostHandler()
        facility_handler = FacilityHandler()

        sheet = workbook.active
        name_handler.set_next(surname_handler).set_next(fathersname_handler).set_next(
            hire_date_handler
        ).set_next(post_handler).set_next(facility_handler)

        if sheet and name_handler.handle(sheet[1]):
            return sheet
        return None

    def read_from_excel(self) -> List[Tuple]:
        name_handler = TypeNameHandler()
        surname_handler = TypeSurnameHandler()
        fathersname_handler = TypeFathresnameHandler()
        hiredate_handler = TypeHireDateHandler()
        facility_handler = TypeFacilityHandler()
        post_handler = TypePostHandler()

        data = []

        try:
            excel_file = str(input("Введите путь до требуемого excel файла: "))
            workbook = load_workbook(excel_file)
        except FileNotFoundError:
            print("Указанный файл не найден")
            input("Нажмите Enter что бы продолжить")
            return data
        except InvalidFileException:
            print("Введён файл не подходящего формата!")
            input("Нажмите Enter что бы продолжить")
            return data

        sheet = self.__check_file(workbook)

        if not sheet:
            return data

        min_row = 2
        for row in sheet.iter_rows(min_row, values_only=True):
            name_handler.set_next(surname_handler).set_next(
                fathersname_handler
            ).set_next(hiredate_handler).set_next(facility_handler).set_next(
                post_handler
            )

            if not name_handler.handle(sheet[min_row]):
                return

            try:
                name = str(row[0])
                surname = str(row[1])
                fathersname = str(row[2])

                hiring_date = row[3]
                post_name = str(row[4])
                facility_name = str(row[5])

                if not isinstance(hiring_date, datetime):
                    hiring_date = datetime.today()

                fetched_post = (
                    self.repository_manager.get_post_repository().get_by_name(post_name)
                )
                if not fetched_post:
                    return data

                fetched_facility = (
                    self.repository_manager.get_facility_repository().get_by_name(
                        facility_name
                    )
                )
                if not fetched_facility:
                    return data

                login = translit(
                    name[:3] + surname[:3] + fathersname[:3], "ru", reversed=True
                )

                if fetched_post.name in UserPost.LEADERSHIP.value:
                    password_length = 12
                else:
                    password_length = 8

                password = PasswordGenerator.generate(password_length)

                data.append(
                    (
                        row[0],
                        row[1],
                        row[2],
                        fetched_facility.name,
                        fetched_post.name,
                        hiring_date,
                        login,
                        password,
                    )
                )
                min_row += 1

            except IndexError:
                print("В таблице недостаточно столбцов")
        return data

    def create_raport(self, user_data: List[Tuple]) -> None:
        if not os.path.exists("./temp/raports"):
            os.mkdir("./temp/raports")

        raportname = str(
            "./temp/raports/raport" + str(datetime.date(datetime.today())) + ".xlsx"
        )

        if not os.path.exists(raportname):
            raport = Workbook()
            raport_sheet = raport.active

            if not raport_sheet:
                return

            raport_sheet["A1"] = "ФИО"
            raport_sheet["B1"] = "Установка"
            raport_sheet["C1"] = "Логин"
            raport_sheet["D1"] = "Пароль"

            raport.save(filename=raportname)

        raport = load_workbook(raportname)
        raport_sheet = raport["Sheet"]

        row_number: int = 2

        for field in user_data:
            raport_sheet[f"A{row_number}"] = f"{field[1]} {field[0]} {field[2]}"
            raport_sheet[f"B{row_number}"] = field[3]
            raport_sheet[f"C{row_number}"] = field[6]
            raport_sheet[f"D{row_number}"] = field[7]
            raport.save(raportname)
            row_number += 1

        raport.close()
