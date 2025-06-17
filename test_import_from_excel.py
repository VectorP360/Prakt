from openpyxl import load_workbook
from psycopg import connect

conn = connect(
    dbname = '',
	user = '',
	password = '',
	host = '',
	port = ''
    )

cursor = conn.cursor()

# путь к excel файлу
excel_file = './data_import.xlsx'

# Присвоение выбранному файлу конкретной переменной
workbook = load_workbook(excel_file)

# Присвоение переменной выбранную страницу из excel файла
sheet = workbook.active

# список из названий столбцов таблицы
column_names = [column.value for column in sheet[1]]

#Список, в котором будут храниться данные
data = []

# Занесение данных в список (На выходе мы получим список data,
# в котором будет множество списков, соответствующих записям)
for row in sheet.iter_rows(min_row=2, values_only=True):
    data.append(row)
    print(row)

for iteration in data:
    cursor.execute("""
   INSERT INTO users (user_id, surname, name, fathersname, facility_id, post_id, hire_date, login, password)
   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
""",(iteration[0],iteration[1],iteration[2],iteration[3],iteration[4],iteration[5],iteration[6],iteration[7],iteration[8]))

conn.commit()