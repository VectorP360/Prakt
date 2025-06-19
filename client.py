from datetime import datetime
import os

from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from transliterate import translit

from repository.repository import RepositoryManager
from schemas.user import UserIn, UserOut
from update_id_to_list import list_for_users, list_for_facilityes, list_for_posts
from password_generator import password_generator
from enum_class import NumChoice , StrChoice
from test_file import read_from_excel, create_raport

class TerminalClient:
    def __init__(self, manager: RepositoryManager): 
        self.facility_repository = manager.get_facility_repository()
        self.posts_repository = manager.get_posts_repository()
        self.user_repository = manager.get_user_repository()
        self.scada_scheme_repository = manager.get_scada_scheme_repository()
        self.facility_types_repository = manager.get_facility_types_repository()
        self.workshop_repository = manager.get_workshop_repository()

    def run(self) -> None:
        operation = None

        while operation != 0:
            print('Укажите номер операции, которую хотите выполнить\n',
                '1: Создание новой записи\n',
                '2: Просмотр записи по ID / всех записей\n',
                '3: Редактирование записи\n',
                '4: удаление записи\n',
                '5: переместить пользователя на другую установку, через excel файл\n',
                '6: Добавление записей из excel файла\n',
                'Для выхода из приложения введие 0\n')

            operation = int(input('Операция :'))

            match operation:
                case NumChoice.create_user:
                    self.add_user()

                case NumChoice.select_user:
                    self.show_users()

                case NumChoice.edit_user:
                    self.update_user()

                case NumChoice.delete_user:
                    self.delete_user()

                case NumChoice.edit_by_excel:
                    self.edit_user_from_excel()

                case NumChoice.create_by_excel:
                    self.add_user_from_excel()

                case NumChoice.exit_programm:
                    return

                case _:
                    print('Выбрана неизвестная операция')
                    return


    def add_user(self):
        
        try:
            print('Для создания записи о сотруднике укажите все перечисленные данные:')
            surname = str(input('Фамилия: '))
            name = str(input('Имя: '))
            fathersname = str(input('Отчество: '))
            
            new_facility = list_for_facilityes(self.facility_repository)
            if not new_facility:
                print('Установки под данным номером не обнаружено')
                input('Нажмите Enter что бы продолжить ')
                return

            new_post = list_for_posts(self.posts_repository)
            if not new_post:
                print('Должности под данным номером не обнаружено')
                input('Нажмите Enter что бы продолжить ')
                return

            new_hire = int(input('\nЭто новый сотрудник, или старый, которого ещё нет в базе данных?\n1: Новый сотрудник\n2: Старый сотрудник\n'))
            match new_hire:

                case NumChoice.new_hire_user:
                    hire_date = datetime.today()

                case NumChoice.old_hire_user:
                    hire_date = datetime.strptime(str(input('Дата найма сотрудника (год-месяц-число): ')), '%y-%m-%d')

            login = translit(name[:3] + surname[:3] + fathersname[:3],"ru", reversed="true")
        except KeyboardInterrupt:
            print('\nВыполнение программы прервано!')
            input('Нажмите Enter что бы продолжить')
            return

        match new_post.name:
            case StrChoice.gendir_post:
                password_length = 12
            case StrChoice.glavspec_post:
                password_length = 12
            case StrChoice.starspec_post:
                password_length = 8
            case StrChoice.spec1categ_post:
                password_length = 8
            case _:
                print('Неизвестная должность')
                return

        password = password_generator(password_length)

        new_user = self.user_repository.create(
            new_user = UserIn(surname, 
                                      name, 
                                      fathersname, 
                                      facility = new_facility, 
                                      post = new_post, 
                                      hire_date = hire_date, 
                                      login = login, 
                                      password = password))

        print(new_user)
            
        input('Нажмите Enter что бы продолжить ')


    def show_users(self):
        # TODO: Описать возможные варианты ответа на данный вопрос как класс, который наследуется от класса IntEnum 
        # (почитай примеры использования).
        # С использованием данного подхода опиши методы, которые будут инкапсулировать логику, описанную после if и elif. [X]
        # Аналогично в других частях кода
        try:
            select_type = int(input('Вы хотите просмотреть одну запись с конкретным ID или сразу все?\n1: C конкретным ID\n2: Все записи\n'))

            match select_type:
                case NumChoice.select_id:
                    needed_id = input('Введите ID записи которую хотите просмотреть\n')
                    founded_user = self.user_repository.get_by_ID(needed_id)
                    
                    if founded_user:
                        print(founded_user)
                    else:
                        print('Записи с указанным ID не найдено')
                
                case NumChoice.select_all_users:
                    for founded_user in (self.user_repository.get_all()):
                        print(founded_user)
                
                case _:
                    print('Выбрана несуществующая операция')
            input('Нажмите Enter что бы продолжить ')
        except KeyboardInterrupt:
            print('Просмотр записи прерван!')
            input('Нажмите Enter что бы продолжить')
            return


    def update_user(self):
        
        try:
            print('Сотрудники: ')
            edit_user = list_for_users(self.user_repository)

            if not edit_user:
                print('Сотрудника под данным номером не обнаружено')
                return
                
            surname = str(input('Новая фамилия: '))
            name = str(input('Новое имя: '))
            fathersname = str(input('Новое отчество: '))
                
            facility_in = list_for_facilityes(self.facility_repository)

            if not facility_in:
                print('Установки под указанным номером не обнаружено')
                return

            post_in = list_for_posts(self.posts_repository)

            if not post_in:
                print('Должности под указанным номером не обнаружено')
                return

            login = str(input('Новый логин сотрудника: '))
            password = edit_user.password

            upd_user = self.user_repository.update(
                edit_user.user_id, 
                new_user = UserIn(
                    surname, 
                    name, 
                    fathersname, 
                    facility_in, 
                    post_in, 
                    edit_user.hire_date, 
                    login, 
                    password)
                )

            print(upd_user)
                    
            input('Нажмите Enter что бы продолжить')
        
        except KeyboardInterrupt:
            print('Редактирование записи прервано!')
            input('Нажмите Enter что бы продолжить')
            return


    def delete_user(self):
        try:
            users_id = []
            print('\nСотрудники:')
            
            for iteration in self.user_repository.get_all():
                print(f'''ID: {iteration.user_id}, ФИО: {iteration.surname} {iteration.name} {iteration.fathersname}, должность: {iteration.post.name}''')
                users_id.append(iteration.user_id)
                
            deleted_user = input('Введите ID сотрудника, чью запись хотите удалить: ')

            if deleted_user not in users_id:
                acception = input('\nВы уверенны, что хотите удалить эту запись? \nПосле удаления её нельзя будет восстановить (y/n): ')

                if acception == 'y':
                                
                    if self.user_repository.delete(deleted_user):
                        print('Удаление произшло успешно!')
                    else:
                        print('Удаление не состоялось!')
                
                else:
                    print('Удаление отменено')
                    
                input('Нажмите Enter что бы продолжить')
        except KeyboardInterrupt:
            print('Удаление записи прервано!')
            input('Нажмите Enter что бы продолжить')
            return
        

    def edit_user_from_excel(self):

        try:
            excel_file = input('Введите путь до требуемого excel файла: ')
            workbook = load_workbook(excel_file)    
        except FileNotFoundError:
            print('Указанный файл не найден')
            input('Нажмите Enter что бы продолжить')
            return
        except InvalidFileException:
            print('Введён файл не подходящего формата!')
            input('Нажмите Enter что бы продолжить')
            return
        
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=1,max_row=1,values_only=True):
            if (row[0] != 'Имя' or 
                row[1] != 'Фамилия' or 
                row[2] != 'Отчество' or 
                row[3] != 'Текущая должность' or 
                row[4] != 'Новая должность' or 
                row[5] != 'Текущая установка' or 
                row[6] != 'Новая установка'):

                return print('Файл не соответствует формату!\n')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                edited_user = self.user_repository.get_by_FIO(
                    name=row[0],
                    surname=row[1],
                    fathersname=row[2],
                    facility_name=row[5],
                    post_name=row[3])

                if row[4]:
                    edited_user.post = self.posts_repository.get_by_name(name=row[4])

                if row[6]:
                    edited_user.facility = self.facility_repository.get_by_name(name=row[6])

                print(self.user_repository.update(
                    edited_user.user_id,
                    new_user = UserIn(
                        edited_user.surname,
                        edited_user.name,
                        edited_user.fathersname,
                        edited_user.facility,
                        edited_user.post,
                        edited_user.hire_date,
                        edited_user.login,
                        edited_user.password)))
                
            except AttributeError:
                pass

    def add_user_from_excel(self):

        data = read_from_excel(self.posts_repository, self.facility_repository)

        if data == None:
            return
        
        print('чтение успешно!')

        # create_raport(data)
        
        # for iteration in data:

        #     fetched_facility = self.facility_repository.get_by_name(iteration[3])
        #     fetched_post = self.posts_repository.get_by_name(iteration[4])

        #     new_user = self.user_repository.create(
        #         new_user = UserIn(
        #             surname = iteration[1], 
        #             name = iteration[0], 
        #             fathersname = iteration[2], 
        #             facility = fetched_facility, 
        #             post = fetched_post, 
        #             hire_date = iteration[5], 
        #             login = iteration[6], 
        #             password = iteration[7]))

        #     print(new_user)
        input('Нажмите Enter что бы продолжить')