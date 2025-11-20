from datetime import datetime
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from transliterate import translit

from repositories.repository import RepositoryManager
from schemas.user import UserIn, UserOut, PostOut, FacilityOut

from tools import PasswordGenerator, NewUsersExcelReader
from enums import Command, UserPost, Acceptance

from handlers.name_handler import NameHandler
from handlers.surname_handler import SurnameHandler
from handlers.fathersname_handler import FathresnameHandler
from handlers.now_post_handler import NowPostHandler
from handlers.new_post_handler import NewPostHandler
from handlers.now_facility_handler import NowFacilityHandler
from handlers.new_facility_handler import NewFacilityHandler


class UserTerminalClient:
    def __init__(self, manager: RepositoryManager): 
        self.facility_repository = manager.get_facility_repository()
        self.post_repository = manager.get_post_repository()
        self.user_repository = manager.get_user_repository()
        self.scada_scheme_repository = manager.get_scada_scheme_repository()
        self.facility_types_repository = manager.get_facility_types_repository()
        self.workshop_repository = manager.get_workshop_repository()
        self.manager = manager
        
    def run(self) -> None:
        operation = None

        while operation != 0:
            print('Укажите номер операции, которую хотите выполнить\n',
                '1: Создание новой записи\n',
                '2: Просмотр записи по ID / всех записей\n',
                '3: Редактирование записи\n',
                '4: удаление записи\n',
                '5: переместить пользователя на другую установку, через excel файл\n',
                '6: Добавление записей из excel файла\n',
                'Для выхода из приложения введие 0\n')

            operation = int(input('Операция :'))

            match operation:
                case Command.CREATE_USER:
                    self.add_user()

                case Command.SELECT_USER:
                    self.show_user()

                case Command.UPDATE_USER:
                    self.update_user()

                case Command.DELETE_USER:
                    self.delete_user()

                case Command.EDIT_BY_EXEL:
                    self.edit_user_from_excel()

                case Command.CREATE_BY_EXCEL:
                    self.add_user_from_excel()

                case Command.EXIT:
                    return

                case _:
                    print('Выбрана неизвестная операция')
                    return


    def add_user(self):
        try:
            print('Для создания записи о сотруднике укажите все перечисленные данные:')
            surname = str(input('Фамилия: '))
            name = str(input('Имя: '))
            fathersname = str(input('Отчество: '))
            
            new_facility = self.select_facility()
            if not new_facility:
                print('Установки под данным номером не обнаружено')
                input('Нажмите Enter что бы продолжить ')
                return

            new_post = self.select_post()
            if not new_post:
                print('Должности под данным номером не обнаружено')
                input('Нажмите Enter что бы продолжить ')
                return

            new_hire = int(input('\nЭто новый сотрудник, или старый, которого ещё нет в базе данных?\n1: Новый сотрудник\n2: Старый сотрудник\n'))
            
            
            if new_hire == Command.NEW_HIRE_USER:
                hire_date = datetime.today()
            else:
                hire_date = datetime.strptime(str(input('Дата найма сотрудника (год-месяц-число): ')), '%y-%m-%d')

            login = translit(name[:3] + surname[:3] + fathersname[:3],"ru", reversed=True)
        
        except KeyboardInterrupt:
            print('\nВыполнение программы прервано!')
            input('Нажмите Enter что бы продолжить')
            return

        
        if new_post.name in UserPost.LEADERSHIP.value:
            password_length = 12
        else:
            password_length = 8

        password = PasswordGenerator.generate(password_length)

        new_user = self.user_repository.create(
            new_user = UserIn(surname, 
                                      name, 
                                      fathersname, 
                                      facility = new_facility, 
                                      post = new_post, 
                                      hire_date = str(hire_date), 
                                      login = login, 
                                      password = password))

        print(new_user)
            
        input('Нажмите Enter что бы продолжить ')


    def show_user(self):
        try:
            select_type = int(input('Вы хотите просмотреть одну запись с конкретным ID или сразу все?\n1: C конкретным ID\n2: Все записи\n'))

            match select_type:
                case Command.SHOW_BY_ID:
                    needed_id = input('Введите ID записи которую хотите просмотреть\n')
                    founded_user = self.user_repository.get_by_ID(needed_id)
                    
                    if founded_user:
                        print(founded_user)
                    else:
                        print('Записи с указанным ID не найдено')
                
                case Command.SHOW_ALL:
                    for founded_user in (self.user_repository.get_all()):
                        print(founded_user)
                
                case _:
                    print('Выбрана несуществующая операция')
            input('Нажмите Enter что бы продолжить ')
        except KeyboardInterrupt:
            print('Просмотр записи прерван!')
            input('Нажмите Enter что бы продолжить')
            return


    def update_user(self):
        try:
            print('Сотрудники: ')
            edit_user = self.select_user()

            if not edit_user:
                print('Сотрудника под данным номером не обнаружено')
                return
                
            surname = str(input('Новая фамилия: '))
            name = str(input('Новое имя: '))
            fathersname = str(input('Новое отчество: '))
                
            facility_in = self.select_facility()

            if not facility_in:
                print('Установки под указанным номером не обнаружено')
                return

            post_in = self.select_post()

            if not post_in:
                print('Должности под указанным номером не обнаружено')
                return

            login = str(input('Новый логин сотрудника: '))
            password = edit_user.password

            upd_user = self.user_repository.update(
                edit_user.user_id, 
                new_user = UserIn(
                    surname, 
                    name, 
                    fathersname, 
                    facility_in, 
                    post_in, 
                    edit_user.hire_date, 
                    login, 
                    password)
                )

            print(upd_user)
                    
            input('Нажмите Enter что бы продолжить')
        
        except KeyboardInterrupt:
            print('Редактирование записи прервано!')
            input('Нажмите Enter что бы продолжить')
            return


    def delete_user(self):
        try:
            user_id = []
            print('\nСотрудники:')
            
            for iteration in self.user_repository.get_all():
                print(f'''ID: {iteration.user_id}, ФИО: {iteration.surname} {iteration.name} {iteration.fathersname}, должность: {iteration.post.name}''')
                user_id.append(iteration.user_id)
                
            deleted_user = int(input('Введите ID сотрудника, чью запись хотите удалить: '))

            if deleted_user in user_id:
                acception = input('\nВы уверенны, что хотите удалить эту запись? \nПосле удаления её нельзя будет восстановить (y/n): ')

                if acception == Acceptance.YES: 
                    if self.user_repository.delete(deleted_user):
                        print('Удаление произшло успешно!')
                    else:
                        print('Удаление не состоялось!')
                
                else:
                    print('Удаление отменено')
                    
                input('Нажмите Enter что бы продолжить')
        except KeyboardInterrupt:
            print('Удаление записи прервано!')
            input('Нажмите Enter что бы продолжить')
            return
        

    def edit_user_from_excel(self):

        name = NameHandler()
        surname = SurnameHandler()
        fathersname = FathresnameHandler()
        now_post = NowPostHandler()
        new_post = NewPostHandler()
        now_facility = NowFacilityHandler()
        new_facility = NewFacilityHandler()

        try:
            excel_file = input('Введите путь до требуемого excel файла: ')
            workbook = load_workbook(excel_file)    
        except FileNotFoundError:
            print('Указанный файл не найден')
            input('Нажмите Enter что бы продолжить')
            return
        except InvalidFileException:
            print('Введён файл не подходящего формата!')
            input('Нажмите Enter что бы продолжить')
            return
        
        sheet = workbook.active
        if not sheet:
            return
        
        name.set_next(surname).set_next(fathersname).set_next(now_post).set_next(new_post).set_next(now_facility).set_next(new_facility)

        row = sheet[1]
        if not row:
            return 
        
        if not name.handle(row):
            return print('Файл не соответствует формату!\n')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                edited_user = self.user_repository.get_by_FIO(
                    name=str(row[0]),
                    surname=str(row[1]),
                    fathersname=str(row[2]),
                    facility_name=str(row[5]),
                    post_name=str(row[3]))

                if not edited_user:
                    return
                
                if row[4]:
                    fetched_post = self.post_repository.get_by_name(name=str(row[4]))
                    if fetched_post:
                        edited_user.post = fetched_post

                if row[6]:
                    fetched_facility = self.facility_repository.get_by_name(name=str(row[6]))
                    if fetched_facility:
                        edited_user.facility = fetched_facility

                print(self.user_repository.update(
                    edited_user.user_id,
                    new_user = UserIn(
                        edited_user.surname,
                        edited_user.name,
                        edited_user.fathersname,
                        edited_user.facility,
                        edited_user.post,
                        edited_user.hire_date,
                        edited_user.login,
                        edited_user.password)))
                
            except AttributeError:
                pass

    def add_user_from_excel(self):
        reader = NewUsersExcelReader(self.manager)
        
        data = reader.read_from_excel()

        if not data:
            return

        reader.create_raport(data)
        
        for user_data in data:

            fetched_facility = self.facility_repository.get_by_name(user_data[3])
            fetched_post = self.post_repository.get_by_name(user_data[4])

            if not fetched_facility:
                print(f"Установка '{fetched_facility}' не была найдена в базе данных")
                return
            
            if not fetched_post:
                print(f"Должность '{fetched_post}' не была найдена в базе данных")
                return
            
            new_user = self.user_repository.create(
                new_user = UserIn(
                    surname = user_data[1], 
                    name = user_data[0], 
                    fathersname = user_data[2], 
                    facility = fetched_facility, 
                    post = fetched_post, 
                    hire_date = user_data[5], 
                    login = user_data[6], 
                    password = user_data[7]))

            print(new_user)
        
        input('Нажмите Enter что бы продолжить')


    def select_user(self) -> Optional[UserOut]:
        user = []
        iteration_number = 0

        for iteration in self.user_repository.get_all():
            print('Номер сотрудника: ',iteration_number,'\n',iteration)
            user.append(iteration.user_id)
            iteration_number += 1

        edit_user = int(input('\nВведите номер сотрудника, чью запись хотите изменить: '))

        try:
            return self.user_repository.get_by_ID(user[edit_user])
        except IndexError:
            return


    def select_facility(self) -> Optional[FacilityOut]:
        facilityes = []
        iteration_number = 0 

        print('\nУстановки:')
        for iteration in self.facility_repository.get_all():
            print(f'Установка №{iteration_number} , {iteration}')
            facilityes.append(iteration.facility_id)
            iteration_number += 1

        facility_id = int(input('На какой установке (номер установки) будет работать сотрудник?: '))

        try:
            return self.facility_repository.get_by_ID(facilityes[facility_id])
        except IndexError:
            return
        

    def select_post(self) -> Optional[PostOut]:
        post = []
        iteration_number = 0

        print('\nДолжности:')
        for iteration in self.post_repository.get_all():
            print(f'Должность №{iteration_number}: Наименование: {iteration.name}')
            post.append(iteration.post_ID)
            iteration_number += 1

        post = int(input('Должность сотрудника (Номер должности): '))

        try:
            return self.post_repository.get_by_ID(post[post])
        except IndexError:
            return
